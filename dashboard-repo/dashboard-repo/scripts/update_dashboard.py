import openpyxl, json, re, sys
from collections import defaultdict, Counter
from pathlib import Path

EXCEL_FILE = Path(__file__).parent.parent / 'Purchasing_management_Rev_3.xlsx'
HTML_FILE  = Path(__file__).parent.parent / 'order_dashboard.html'

print(f"Reading: {EXCEL_FILE}")
wb = openpyxl.load_workbook(EXCEL_FILE, data_only=True)
ws = wb['📋 ORDERS']

rows = []
for r in range(4, 700):
    skf = ws.cell(row=r, column=3).value
    if skf is None:
        continue
    tata    = ws.cell(row=r, column=4).value
    vulcan  = ws.cell(row=r, column=5).value
    odate   = ws.cell(row=r, column=9).value
    wh      = ws.cell(row=r, column=11).value
    qo      = ws.cell(row=r, column=12).value
    qs      = ws.cell(row=r, column=13).value
    invoice = ws.cell(row=r, column=17).value
    status  = ws.cell(row=r, column=18).value
    qp      = ws.cell(row=r, column=21).value
    deliv   = ws.cell(row=r, column=24).value

    inv_str = str(invoice).strip() if invoice else ''
    if inv_str == '0000000':
        inv_str = ''

    rows.append({
        'skf':    str(skf).strip(),
        'tata':   str(tata).strip() if tata else '',
        'vulcan': str(vulcan).strip() if vulcan else '',
        'date':   odate.strftime('%Y-%m') if hasattr(odate, 'strftime') else '',
        'wh':     str(wh).strip() if wh else 'N/A',
        'qo':     int(qo) if qo else 0,
        'qs':     int(qs) if qs else 0,
        'qp':     int(qp) if qp else 0,
        'inv':    inv_str,
        'st':     str(status).strip() if status else 'PENDING',
        'del':    1 if str(deliv).strip() == '✔' else 0,
    })

monthly = defaultdict(lambda: [0, 0])
for r in rows:
    if r['date']:
        monthly[r['date']][0] += r['qo']
        monthly[r['date']][1] += r['qs']

skf_map = defaultdict(lambda: {'lines':0,'qo':0,'qs':0,'sts':[],'invs':set(),'whs':set(),'tatas':set(),'vulcans':set(),'date':''})
for r in rows:
    k = r['skf']
    skf_map[k]['lines'] += 1
    skf_map[k]['qo']    += r['qo']
    skf_map[k]['qs']    += r['qs']
    skf_map[k]['sts'].append(r['st'])
    if r['inv']:    skf_map[k]['invs'].add(r['inv'])
    skf_map[k]['whs'].add(r['wh'])
    if r['tata']:   skf_map[k]['tatas'].add(r['tata'])
    if r['vulcan']: skf_map[k]['vulcans'].add(r['vulcan'])
    if r['date'] and not skf_map[k]['date']:
        skf_map[k]['date'] = r['date']

skf_list = []
for skf, d in skf_map.items():
    dom_st = Counter(d['sts']).most_common(1)[0][0]
    skf_list.append({
        'skf': skf, 'lines': d['lines'], 'qo': d['qo'], 'qs': d['qs'],
        'st': dom_st, 'invs': len(d['invs']), 'whs': list(d['whs']),
        'tatas': list(d['tatas']), 'vulcans': list(d['vulcans']), 'date': d['date']
    })
skf_list.sort(key=lambda x: x['date'])

tata_map = defaultdict(lambda: {'lines':0,'qo':0,'qs':0,'sts':[],'skfs':set()})
for r in rows:
    if not r['tata']: continue
    tata_map[r['tata']]['lines'] += 1
    tata_map[r['tata']]['qo']   += r['qo']
    tata_map[r['tata']]['qs']   += r['qs']
    tata_map[r['tata']]['sts'].append(r['st'])
    tata_map[r['tata']]['skfs'].add(r['skf'])

vulcan_map = defaultdict(lambda: {'lines':0,'qo':0,'qs':0,'sts':[],'skfs':set()})
for r in rows:
    if not r['vulcan']: continue
    vulcan_map[r['vulcan']]['lines'] += 1
    vulcan_map[r['vulcan']]['qo']   += r['qo']
    vulcan_map[r['vulcan']]['qs']   += r['qs']
    vulcan_map[r['vulcan']]['sts'].append(r['st'])
    vulcan_map[r['vulcan']]['skfs'].add(r['skf'])

status_counts = Counter(r['st'] for r in rows)
wh_counts     = Counter(r['wh'] for r in rows)

agg = {
    'rows':          rows,
    'status_counts': dict(status_counts),
    'wh_counts':     dict(wh_counts),
    'months':        [[m, v[0], v[1]] for m, v in sorted(monthly.items())],
    'skf_list':      skf_list,
    'tata_list':     [{'tata':k,'lines':v['lines'],'qo':v['qo'],'qs':v['qs'],
                       'st':Counter(v['sts']).most_common(1)[0][0],'skfs':len(v['skfs'])}
                      for k, v in sorted(tata_map.items())],
    'vulcan_list':   [{'vulcan':k,'lines':v['lines'],'qo':v['qo'],'qs':v['qs'],
                       'st':Counter(v['sts']).most_common(1)[0][0],'skfs':len(v['skfs'])}
                      for k, v in sorted(vulcan_map.items())],
    'total_lines':    len(rows),
    'total_qo':       sum(r['qo'] for r in rows),
    'total_qs':       sum(r['qs'] for r in rows),
    'total_invoices': len(set(r['inv'] for r in rows if r['inv'])),
    'delivered':      sum(r['del'] for r in rows),
}

new_data = json.dumps(agg, separators=(',', ':'))

html = HTML_FILE.read_text(encoding='utf-8')
new_html = re.sub(
    r'const D = \{.*?\};',
    f'const D = {new_data};',
    html,
    count=1,
    flags=re.DOTALL
)

if new_html == html:
    print("ERROR: data block not found in HTML — check the pattern.")
    sys.exit(1)

HTML_FILE.write_text(new_html, encoding='utf-8')
print(f"Done — {agg['total_lines']} lines, {agg['total_qo']} ordered, {agg['total_qs']} shipped")
print(f"Status breakdown: {dict(status_counts)}")
